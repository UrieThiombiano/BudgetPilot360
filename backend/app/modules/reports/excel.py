"""
Rendu Excel du rapport financier (openpyxl). Cinq feuilles formatées :
Résumé (Recettes | Dépenses | Bénéfice), Recettes (détail), Dépenses (détail),
Recettes par catégorie, Dépenses par catégorie.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INDIGO = "4F46E5"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=INDIGO)
TITLE_FONT = Font(bold=True, size=14, color="1C1C28")
LABEL_FONT = Font(bold=True, size=10, color="52514E")
THIN_BORDER = Border(bottom=Side(style="thin", color="E4E4EC"))
MONEY_FORMAT = '#,##0.00 "FCFA"'
PCT_FORMAT = "0 %"
GREEN_FONT = Font(bold=True, size=10, color="116329")
RED_FONT = Font(bold=True, size=10, color="A12622")

STATUS_FILLS = {
    "approved": PatternFill("solid", fgColor="E7F6EC"),
    "pending": PatternFill("solid", fgColor="FDF3DC"),
    "rejected": PatternFill("solid", fgColor="FDEAEA"),
}


def _header_row(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


def _fit_columns(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _fr_date(d) -> str:
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def _line(ws, row, label, amount, count=None, *, font=None):
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    amount_cell = ws.cell(row=row, column=2, value=amount)
    amount_cell.number_format = MONEY_FORMAT
    if font is not None:
        amount_cell.font = font
    if count is not None:
        ws.cell(row=row, column=3, value=f"{count} ligne(s)").font = Font(size=9, color="6B6B76")


def _detail_sheet(wb, title, rows, headers, *, with_source, widths):
    ws = wb.create_sheet(title)
    _header_row(ws, 1, headers)
    for r, x in enumerate(rows, start=2):
        col = 1
        ws.cell(row=r, column=col, value=str(x["date"])[:10]); col += 1
        ws.cell(row=r, column=col, value=x["category_name"]); col += 1
        ws.cell(row=r, column=col, value=x["author_name"]); col += 1
        if with_source:
            ws.cell(row=r, column=col, value=x.get("source", "")); col += 1
        ws.cell(row=r, column=col, value=x["description"]); col += 1
        amount_cell = ws.cell(row=r, column=col, value=x["amount"])
        amount_cell.number_format = MONEY_FORMAT
        col += 1
        status_cell = ws.cell(row=r, column=col, value=x["status_label"])
        status_cell.fill = STATUS_FILLS[x["status"]]
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
    _fit_columns(ws, widths)


def _breakdown_sheet(wb, title, items, headers, widths):
    ws = wb.create_sheet(title)
    _header_row(ws, 1, headers)
    for r, c in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=c["name"])
        planned = ws.cell(row=r, column=2, value=c["planned_budget"])
        planned.number_format = MONEY_FORMAT
        consumed = ws.cell(row=r, column=3, value=c["consumed"])
        consumed.number_format = MONEY_FORMAT
        if c["ratio"] is not None:
            ratio = ws.cell(row=r, column=4, value=c["ratio"])
            ratio.number_format = PCT_FORMAT
            if c["ratio"] > 1:
                ratio.font = Font(bold=True, color="A12622")
        else:
            ws.cell(row=r, column=4, value="—")
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = THIN_BORDER
    ws.freeze_panes = "A2"
    _fit_columns(ws, widths)


def render_excel(data: dict) -> bytes:
    wb = Workbook()

    # --- Feuille Résumé : Recettes | Dépenses | Bénéfice ---
    ws = wb.active
    ws.title = "Résumé"
    ws["A1"] = f"Rapport financier — {data['company_name']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Période du {_fr_date(data['date_from'])} au {_fr_date(data['date_to'])} · "
        f"généré le {_fr_date(data['generated_on'])} · BudgetPilot360"
    )
    ws["A2"].font = Font(size=9, color="6B6B76")

    net_font = GREEN_FONT if data["net_profit"] >= 0 else RED_FONT
    _line(ws, 4, "Recettes confirmées", data["total_revenue"], data["count_revenue_approved"])
    _line(ws, 5, "Dépenses approuvées", data["total_approved"], data["count_approved"])
    _line(ws, 6, "Bénéfice net", data["net_profit"], font=net_font)
    ws.cell(row=7, column=1, value="Marge").font = LABEL_FONT
    if data["margin"] is not None:
        margin_cell = ws.cell(row=7, column=2, value=data["margin"] / 100)
        margin_cell.number_format = PCT_FORMAT
        margin_cell.font = net_font
    else:
        ws.cell(row=7, column=2, value="—")

    _line(ws, 9, "Budget annuel", data["annual_budget"])
    _line(ws, 10, "Dépenses en attente", data["total_pending"], data["count_pending"])
    _line(ws, 11, "Dépenses rejetées", data["total_rejected"], data["count_rejected"])
    _line(ws, 12, "Recettes en attente", data["total_revenue_pending"], data["count_revenue_pending"])
    _fit_columns(ws, [30, 18, 16])

    # --- Détails ---
    _detail_sheet(
        wb, "Recettes", data["revenues"],
        ["Date", "Catégorie", "Auteur", "Source / client", "Description", "Montant", "Statut"],
        with_source=True, widths=[12, 20, 22, 24, 34, 14, 12],
    )
    _detail_sheet(
        wb, "Dépenses", data["expenses"],
        ["Date", "Catégorie", "Auteur", "Description", "Montant", "Statut"],
        with_source=False, widths=[12, 20, 26, 42, 14, 12],
    )

    # --- Répartitions par catégorie ---
    _breakdown_sheet(
        wb, "Recettes par catégorie", data["revenue_breakdown"],
        ["Catégorie", "Objectif", "Réalisé (période)", "% objectif"], [26, 16, 20, 14],
    )
    _breakdown_sheet(
        wb, "Dépenses par catégorie", data["breakdown"],
        ["Catégorie", "Budget prévu", "Consommé (période)", "% consommé"], [26, 16, 20, 14],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
