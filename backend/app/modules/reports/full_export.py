"""
Export complet des données d'une entreprise (portabilité) — un classeur Excel,
une feuille par famille : Entreprise, Équipe, Catégories, Dépenses, Recettes,
Automatisations, Journal d'audit.

Réservé aux admins de l'entreprise (adjoint compris), audité côté router :
c'est une sortie de données hors de la plateforme. Les justificatifs (fichiers
Storage) ne sont pas embarqués dans le classeur — ils restent téléchargeables
un à un depuis l'app ; la colonne « Justificatif » indique leur présence.

Styles réutilisés depuis reports/excel.py (même identité que les rapports).
"""

import io
import json
from datetime import datetime, timezone

from openpyxl import Workbook

from app.core.supabase_client import get_service_client
from app.modules.reports.excel import (
    LABEL_FONT,
    MONEY_FORMAT,
    STATUS_FILLS,
    THIN_BORDER,
    TITLE_FONT,
    _fit_columns,
    _header_row,
)

EXPENSE_STATUS_LABELS = {"approved": "Approuvée", "pending": "En attente", "rejected": "Rejetée"}
REVENUE_STATUS_LABELS = {"approved": "Confirmée", "pending": "En attente", "rejected": "Rejetée"}
SUBSCRIPTION_LABELS = {"active": "Actif", "suspended": "Suspendu"}
AUDIT_MAX_ROWS = 5000


def _dt(value) -> str:
    """timestamptz ISO → « JJ/MM/AAAA HH:MM » (chaîne vide si absent)."""
    if not value:
        return ""
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return str(value)


def _d(value) -> str:
    return str(value)[:10] if value else ""


def _fetch(client, company_id: str) -> dict:
    company = (
        client.table("companies")
        .select("name, annual_budget, plan, subscription_status, subscription_ends_at, created_at, owner_id")
        .eq("id", company_id)
        .execute()
    ).data
    profiles = (
        client.table("profiles")
        .select("id, full_name, email, job_title, role, removed_at, created_at")
        .eq("company_id", company_id)
        .execute()
    ).data or []
    categories = (
        client.table("categories")
        .select("id, name, type, planned_budget, created_at")
        .eq("company_id", company_id)
        .execute()
    ).data or []
    expenses = (
        client.table("expenses")
        .select(
            "expense_date, amount, description, status, rejection_reason, "
            "receipt_path, recurring_id, created_at, category_id, user_id, reviewed_by"
        )
        .eq("company_id", company_id)
        .order("expense_date", desc=True)
        .execute()
    ).data or []
    revenues = (
        client.table("revenues")
        .select(
            "revenue_date, amount, description, source, status, recurring_id, "
            "proof_path, created_at, category_id, user_id"
        )
        .eq("company_id", company_id)
        .order("revenue_date", desc=True)
        .execute()
    ).data or []
    recurring = {}
    for kind, table in (("expense", "recurring_expenses"), ("revenue", "recurring_revenues")):
        recurring[kind] = (
            client.table(table)
            .select(
                "description, amount, day_of_month, months_total, months_done, "
                "active, next_due, category_id, created_at"
            )
            .eq("company_id", company_id)
            .order("created_at", desc=True)
            .execute()
        ).data or []
    audit_logs = (
        client.table("audit_logs")
        .select("created_at, action, actor_id, details")
        .eq("company_id", company_id)
        .order("created_at", desc=True)
        .limit(AUDIT_MAX_ROWS)
        .execute()
    ).data or []

    return {
        "company": company[0] if company else {},
        "profiles": profiles,
        "categories": categories,
        "expenses": expenses,
        "revenues": revenues,
        "recurring": recurring,
        "audit_logs": audit_logs,
    }


def _company_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Entreprise"
    company = data["company"]
    ws["A1"] = f"BudgetPilot360 — Export des données · {company.get('name', '')}"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Entreprise", company.get("name", "")),
        ("Budget annuel", float(company.get("annual_budget") or 0)),
        ("Offre", (company.get("plan") or "starter").capitalize()),
        ("Abonnement", SUBSCRIPTION_LABELS.get(company.get("subscription_status") or "active", "")),
        ("Abonnement jusqu'au", _d(company.get("subscription_ends_at"))),
        ("Cliente depuis le", _dt(company.get("created_at"))),
        ("Membres (actifs / total)", ""),
        ("Export généré le", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
    ]
    actifs = sum(1 for p in data["profiles"] if p.get("removed_at") is None)
    rows[6] = ("Membres (actifs / total)", f"{actifs} / {len(data['profiles'])}")

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=i, column=2, value=value)
        if label == "Budget annuel":
            cell.number_format = MONEY_FORMAT

    counts = [
        ("Dépenses", len(data["expenses"])),
        ("Recettes", len(data["revenues"])),
        ("Catégories", len(data["categories"])),
        ("Automatisations", len(data["recurring"]["expense"]) + len(data["recurring"]["revenue"])),
        ("Lignes d'audit exportées", len(data["audit_logs"])),
    ]
    start = len(rows) + 4
    ws.cell(row=start, column=1, value="Contenu du classeur").font = LABEL_FONT
    for i, (label, count) in enumerate(counts, start=start + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=count)
    _fit_columns(ws, [28, 34])


def _team_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Équipe")
    _header_row(ws, 1, ["Nom", "Email", "Fonction", "Rôle", "Statut", "Membre depuis"])
    owner_id = data["company"].get("owner_id")
    members = [p for p in data["profiles"] if p.get("role") != "super_admin"]
    for r, p in enumerate(members, start=2):
        if p.get("role") == "admin":
            role = "Admin principal" if p["id"] == owner_id else "Admin adjoint"
        else:
            role = "Collaborateur"
        values = [
            p.get("full_name") or "",
            p.get("email") or "",
            p.get("job_title") or "",
            role,
            "Actif" if p.get("removed_at") is None else "Retiré",
            _dt(p.get("created_at")),
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v).border = THIN_BORDER
    ws.freeze_panes = "A2"
    _fit_columns(ws, [24, 32, 22, 16, 10, 18])


def _categories_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Catégories")
    _header_row(ws, 1, ["Nom", "Type", "Budget prévu", "Créée le"])
    for r, c in enumerate(data["categories"], start=2):
        ws.cell(row=r, column=1, value=c["name"]).border = THIN_BORDER
        kind = ws.cell(row=r, column=2, value="Recette" if c.get("type") == "revenue" else "Dépense")
        kind.border = THIN_BORDER
        planned = ws.cell(row=r, column=3, value=float(c.get("planned_budget") or 0))
        planned.number_format = MONEY_FORMAT
        planned.border = THIN_BORDER
        ws.cell(row=r, column=4, value=_dt(c.get("created_at"))).border = THIN_BORDER
    ws.freeze_panes = "A2"
    _fit_columns(ws, [28, 10, 18, 18])


def _tx_sheet(wb: Workbook, *, title, rows, names, cat_names, date_key, status_labels, with_source, proof_key):
    headers = ["Date", "Catégorie", "Auteur"]
    if with_source:
        headers.append("Source / client")
    headers += ["Description", "Montant", "Statut", "Motif de rejet", "Justificatif", "Automatique", "Validée par", "Saisie le"]
    ws = wb.create_sheet(title)
    _header_row(ws, 1, headers)
    for r, x in enumerate(rows, start=2):
        values = [_d(x.get(date_key)), cat_names.get(x.get("category_id"), ""), names.get(x.get("user_id"), "")]
        if with_source:
            values.append(x.get("source") or "")
        values += [
            x.get("description") or "",
            float(x.get("amount") or 0),
            status_labels.get(x.get("status"), x.get("status")),
            x.get("rejection_reason") or "",
            "Oui" if x.get(proof_key) else "Non",
            "Oui" if x.get("recurring_id") else "Non",
            names.get(x.get("reviewed_by"), ""),
            _dt(x.get("created_at")),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN_BORDER
            if headers[c - 1] == "Montant":
                cell.number_format = MONEY_FORMAT
            if headers[c - 1] == "Statut" and x.get("status") in STATUS_FILLS:
                cell.fill = STATUS_FILLS[x["status"]]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _fit_columns(ws, [11, 20, 20] + ([20] if with_source else []) + [34, 16, 12, 26, 11, 11, 20, 17])


def _recurring_sheet(wb: Workbook, data: dict, cat_names: dict) -> None:
    ws = wb.create_sheet("Automatisations")
    _header_row(
        ws, 1,
        ["Type", "Libellé", "Catégorie", "Montant mensuel", "Jour du mois",
         "Progression", "Statut", "Prochaine échéance", "Créée le"],
    )
    r = 2
    for kind, label in (("expense", "Dépense"), ("revenue", "Recette")):
        for a in data["recurring"][kind]:
            finished = a["months_done"] >= a["months_total"]
            statut = "Terminée" if finished else ("Active" if a.get("active") else "En pause")
            values = [
                label,
                a.get("description") or "",
                cat_names.get(a.get("category_id"), ""),
                float(a.get("amount") or 0),
                a.get("day_of_month"),
                f"{a['months_done']}/{a['months_total']}",
                statut,
                "" if finished else _d(a.get("next_due")),
                _dt(a.get("created_at")),
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = THIN_BORDER
                if c == 4:
                    cell.number_format = MONEY_FORMAT
            r += 1
    ws.freeze_panes = "A2"
    _fit_columns(ws, [10, 30, 20, 17, 12, 12, 10, 18, 17])


def _audit_sheet(wb: Workbook, data: dict, names: dict) -> None:
    ws = wb.create_sheet("Journal d'audit")
    _header_row(ws, 1, ["Date", "Action", "Acteur", "Détails"])
    for r, log in enumerate(data["audit_logs"], start=2):
        details = log.get("details")
        detail_text = (
            json.dumps(details, ensure_ascii=False, default=str) if isinstance(details, (dict, list)) else (details or "")
        )
        values = [
            _dt(log.get("created_at")),
            log.get("action") or "",
            names.get(log.get("actor_id"), ""),
            detail_text,
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v).border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _fit_columns(ws, [17, 28, 22, 70])


def render_company_export(company_id: str) -> bytes:
    """Construit le classeur complet des données de l'entreprise."""
    data = _fetch(get_service_client(), company_id)
    names = {p["id"]: (p.get("full_name") or p.get("email") or "") for p in data["profiles"]}
    cat_names = {c["id"]: c["name"] for c in data["categories"]}

    wb = Workbook()
    _company_sheet(wb, data)
    _team_sheet(wb, data)
    _categories_sheet(wb, data)
    _tx_sheet(
        wb, title="Dépenses", rows=data["expenses"], names=names, cat_names=cat_names,
        date_key="expense_date", status_labels=EXPENSE_STATUS_LABELS, with_source=False,
        proof_key="receipt_path",
    )
    _tx_sheet(
        wb, title="Recettes", rows=data["revenues"], names=names, cat_names=cat_names,
        date_key="revenue_date", status_labels=REVENUE_STATUS_LABELS, with_source=True,
        proof_key="proof_path",
    )
    _recurring_sheet(wb, data, cat_names)
    _audit_sheet(wb, data, names)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
