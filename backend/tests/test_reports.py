"""Tests des rapports (Phase 7.1) : données, PDF, Excel, RBAC, période, audit.
Inclut l'export complet des données de l'entreprise (portabilité)."""

import io
from types import SimpleNamespace
from unittest.mock import MagicMock

from openpyxl import load_workbook

import app.modules.reports.full_export as full_export
import app.modules.reports.service as reports_service

COMPANY = {
    "id": "11111111-1111-1111-1111-111111111111",
    "name": "Acme",
    "annual_budget": "50000.00",
}
CATEGORIES = [
    {"id": "c1", "name": "Transport", "type": "expense", "planned_budget": "1000.00"},
    {"id": "c2", "name": "Salaires", "type": "expense", "planned_budget": "0.00"},
    {"id": "cr1", "name": "Ventes", "type": "revenue", "planned_budget": "2000.00"},
]
PROFILES = [
    {"id": "u1", "full_name": "Jean User", "email": "jean@acme-corp.fr"},
]
EXPENSES = [
    {"id": "e1", "amount": "100.00", "expense_date": "2026-07-02", "description": "Taxi",
     "status": "approved", "category_id": "c1", "user_id": "u1"},
    {"id": "e2", "amount": "50.00", "expense_date": "2026-07-05", "description": "Prime",
     "status": "approved", "category_id": "c2", "user_id": "u1"},
    {"id": "e3", "amount": "40.00", "expense_date": "2026-07-08", "description": "Péage",
     "status": "pending", "category_id": "c1", "user_id": "u1"},
    {"id": "e4", "amount": "60.00", "expense_date": "2026-07-09", "description": "Resto",
     "status": "rejected", "category_id": "c1", "user_id": "u1"},
]
REVENUES = [
    {"id": "r1", "amount": "400.00", "revenue_date": "2026-07-03", "description": "Vente comptant",
     "status": "approved", "category_id": "cr1", "user_id": "u1", "source": "Client A"},
    {"id": "r2", "amount": "100.00", "revenue_date": "2026-07-06", "description": "Acompte",
     "status": "pending", "category_id": "cr1", "user_id": "u1", "source": "Client B"},
]

PERIOD = "date_from=2026-07-01&date_to=2026-07-31"


def _mock_reports_client(monkeypatch, expenses=None, revenues=None):
    mock_client = MagicMock()
    tables = {
        "companies": MagicMock(),
        "categories": MagicMock(),
        "profiles": MagicMock(),
        "expenses": MagicMock(),
        "revenues": MagicMock(),
    }
    mock_client.table.side_effect = lambda name: tables[name]
    mock_client.tables = tables

    tables["companies"].select.return_value.eq.return_value.execute.return_value = SimpleNamespace(
        data=[COMPANY]
    )
    tables["categories"].select.return_value.eq.return_value.execute.return_value = SimpleNamespace(
        data=CATEGORIES
    )
    tables["profiles"].select.return_value.eq.return_value.execute.return_value = SimpleNamespace(
        data=PROFILES
    )
    tables["expenses"].select.return_value.eq.return_value.gte.return_value.lte.return_value.order.return_value.execute.return_value = SimpleNamespace(
        data=EXPENSES if expenses is None else expenses
    )
    tables["revenues"].select.return_value.eq.return_value.gte.return_value.lte.return_value.order.return_value.execute.return_value = SimpleNamespace(
        data=REVENUES if revenues is None else revenues
    )

    monkeypatch.setattr(reports_service, "get_service_client", lambda: mock_client)
    return mock_client


def test_export_excel(client, as_admin, monkeypatch, silence_audit):
    _mock_reports_client(monkeypatch)

    resp = client.get(f"/reports/export?format=excel&{PERIOD}")

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert 'filename="rapport_budgetpilot360_2026-07-01_2026-07-31.xlsx"' in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == [
        "Résumé", "Graphiques", "Recettes", "Dépenses",
        "Recettes par catégorie", "Dépenses par catégorie",
    ]

    resume = wb["Résumé"]
    assert "Acme" in resume["A1"].value
    assert resume["B4"].value == 400.0  # recettes confirmées
    assert resume["B5"].value == 150.0  # dépenses approuvées : 100 + 50
    assert resume["B6"].value == 250.0  # bénéfice net = 400 - 150
    assert resume["B10"].value == 40.0  # dépenses en attente
    assert resume["B11"].value == 60.0  # dépenses rejetées

    recettes = wb["Recettes"]
    assert recettes.max_row == 3  # entête + 2 recettes
    assert recettes["D2"].value == "Client A"  # colonne Source / client
    assert recettes["F2"].value == 400.0
    assert recettes["G2"].value == "Confirmée"

    depenses = wb["Dépenses"]
    assert depenses.max_row == 5  # entête + 4 dépenses
    assert depenses["E2"].value == 100.0
    assert depenses["F2"].value == "Approuvée"

    rev_cats = wb["Recettes par catégorie"]
    assert rev_cats["A2"].value == "Ventes"
    assert rev_cats["C2"].value == 400.0  # réalisé (confirmées)

    exp_cats = wb["Dépenses par catégorie"]
    assert exp_cats["A2"].value == "Transport"  # trié par consommé décroissant
    assert exp_cats["C2"].value == 100.0  # seules les approuvées comptent
    assert exp_cats["D3"].value == "—"  # budget prévu à 0 → pas de ratio


def test_export_excel_summary_scope(client, as_admin, monkeypatch, silence_audit):
    """`scope=summary` : uniquement Résumé + Graphiques (bilan pour lecteur pressé)."""
    _mock_reports_client(monkeypatch)

    resp = client.get(f"/reports/export?format=excel&scope=summary&{PERIOD}")

    assert resp.status_code == 200
    assert "_resume.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Résumé", "Graphiques"]


def test_report_data_endpoint(client, as_admin, monkeypatch, silence_audit):
    """GET /reports/data : la source unique du flux Générer → Aperçu → Décider."""
    _mock_reports_client(monkeypatch)

    resp = client.get(f"/reports/data?{PERIOD}")

    assert resp.status_code == 200
    body = resp.json()
    assert body["company_name"] == "Acme"
    assert body["total_revenue"] == 400.0
    assert body["total_approved"] == 150.0
    assert body["net_profit"] == 250.0
    # Répartitions avec id + count (mêmes couleurs par catégorie côté front)
    transport = next(c for c in body["breakdown"] if c["name"] == "Transport")
    assert transport["id"] == "c1" and transport["count"] == 1
    # Série mensuelle de la période (zone chart de l'aperçu)
    assert body["monthly"] == [
        {"month": "2026-07", "revenues": 400.0, "expenses": 150.0, "net": 250.0}
    ]
    assert any(c["action"] == "report.generated" for c in silence_audit)


def test_report_data_forbidden_for_user(client, as_user, monkeypatch):
    mock_client = _mock_reports_client(monkeypatch)
    assert client.get(f"/reports/data?{PERIOD}").status_code == 403
    mock_client.table.assert_not_called()


def test_export_pdf(client, as_admin, monkeypatch, silence_audit):
    _mock_reports_client(monkeypatch)

    resp = client.get(f"/reports/export?format=pdf&{PERIOD}")

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"
    assert len(resp.content) > 2000


def test_export_scopes_period_and_company(client, as_admin, monkeypatch, silence_audit):
    mock_client = _mock_reports_client(monkeypatch)

    client.get(f"/reports/export?format=excel&{PERIOD}")

    chain = mock_client.tables["expenses"].select.return_value
    assert chain.eq.call_args.args == ("company_id", as_admin.company_id)
    assert chain.eq.return_value.gte.call_args.args == ("expense_date", "2026-07-01")
    assert chain.eq.return_value.gte.return_value.lte.call_args.args == ("expense_date", "2026-07-31")


def test_export_is_audited(client, as_admin, monkeypatch, silence_audit):
    _mock_reports_client(monkeypatch)

    client.get(f"/reports/export?format=excel&{PERIOD}")

    entry = next(c for c in silence_audit if c["action"] == "report.exported")
    assert entry["details"]["format"] == "excel"
    assert entry["details"]["date_from"] == "2026-07-01"


def test_export_invalid_period(client, as_admin, monkeypatch):
    _mock_reports_client(monkeypatch)

    resp = client.get("/reports/export?format=pdf&date_from=2026-07-31&date_to=2026-07-01")

    assert resp.status_code == 422


def test_export_invalid_format(client, as_admin, monkeypatch):
    _mock_reports_client(monkeypatch)

    resp = client.get(f"/reports/export?format=csv&{PERIOD}")

    assert resp.status_code == 422


def test_export_forbidden_for_user(client, as_user, monkeypatch):
    mock_client = _mock_reports_client(monkeypatch)

    resp = client.get(f"/reports/export?format=pdf&{PERIOD}")

    assert resp.status_code == 403
    mock_client.table.assert_not_called()


def test_export_empty_period(client, as_admin, monkeypatch, silence_audit):
    _mock_reports_client(monkeypatch, expenses=[], revenues=[])

    resp = client.get(f"/reports/export?format=excel&{PERIOD}")

    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb["Résumé"]["B4"].value == 0.0  # recettes
    assert wb["Résumé"]["B5"].value == 0.0  # dépenses
    assert wb["Dépenses"].max_row >= 1


# --- Export complet des données de l'entreprise (portabilité) ---

FULL_COMPANY = {
    **COMPANY,
    "plan": "standard",
    "subscription_status": "active",
    "subscription_ends_at": "2027-07-17",
    "created_at": "2026-07-01T10:00:00+00:00",
    "owner_id": "a1",
}
FULL_PROFILES = [
    {"id": "a1", "full_name": "Awa Admin", "email": "awa@acme-corp.fr", "job_title": "DG",
     "role": "admin", "removed_at": None, "created_at": "2026-07-01T10:00:00+00:00"},
    {"id": "u1", "full_name": "Jean User", "email": "jean@acme-corp.fr", "job_title": "Commercial",
     "role": "user", "removed_at": "2026-07-10T00:00:00+00:00", "created_at": "2026-07-02T10:00:00+00:00"},
]
FULL_EXPENSES = [
    {"expense_date": "2026-07-02", "amount": "100.00", "description": "Taxi", "status": "approved",
     "rejection_reason": None, "receipt_path": "co/e1/x.pdf", "recurring_id": None,
     "created_at": "2026-07-02T09:00:00+00:00", "category_id": "c1", "user_id": "u1", "reviewed_by": "a1"},
]
FULL_REVENUES = [
    {"revenue_date": "2026-07-03", "amount": "400.00", "description": "Vente", "status": "approved",
     "source": "Client A", "recurring_id": "rr1", "proof_path": None,
     "created_at": "2026-07-03T09:00:00+00:00", "category_id": "cr1", "user_id": "a1"},
]
FULL_RECURRING = [
    {"description": "Licence compta", "amount": "45000.00", "day_of_month": 1, "months_total": 12,
     "months_done": 2, "active": True, "next_due": "2026-08-01", "category_id": "c1",
     "created_at": "2026-07-01T10:00:00+00:00"},
]
FULL_AUDIT = [
    {"created_at": "2026-07-05T08:00:00+00:00", "action": "expense.approved", "actor_id": "a1",
     "details": {"id": "e1", "amount": "100.00"}},
]


def _mock_full_export_client(monkeypatch):
    mock_client = MagicMock()
    tables = {
        "companies": MagicMock(),
        "profiles": MagicMock(),
        "categories": MagicMock(),
        "expenses": MagicMock(),
        "revenues": MagicMock(),
        "recurring_expenses": MagicMock(),
        "recurring_revenues": MagicMock(),
        "audit_logs": MagicMock(),
    }
    mock_client.table.side_effect = lambda name: tables[name]
    mock_client.tables = tables

    tables["companies"].select.return_value.eq.return_value.execute.return_value = SimpleNamespace(
        data=[FULL_COMPANY]
    )
    tables["profiles"].select.return_value.eq.return_value.execute.return_value = SimpleNamespace(
        data=FULL_PROFILES
    )
    tables["categories"].select.return_value.eq.return_value.execute.return_value = SimpleNamespace(
        data=CATEGORIES
    )
    tables["expenses"].select.return_value.eq.return_value.order.return_value.execute.return_value = SimpleNamespace(
        data=FULL_EXPENSES
    )
    tables["revenues"].select.return_value.eq.return_value.order.return_value.execute.return_value = SimpleNamespace(
        data=FULL_REVENUES
    )
    tables["recurring_expenses"].select.return_value.eq.return_value.order.return_value.execute.return_value = SimpleNamespace(
        data=FULL_RECURRING
    )
    tables["recurring_revenues"].select.return_value.eq.return_value.order.return_value.execute.return_value = SimpleNamespace(
        data=[]
    )
    tables["audit_logs"].select.return_value.eq.return_value.order.return_value.limit.return_value.execute.return_value = SimpleNamespace(
        data=FULL_AUDIT
    )

    monkeypatch.setattr(full_export, "get_service_client", lambda: mock_client)
    return mock_client


def test_company_export_full_workbook(client, as_admin, monkeypatch, silence_audit):
    """Export complet : une feuille par famille, tous statuts et toutes périodes."""
    _mock_full_export_client(monkeypatch)

    resp = client.get("/reports/company-export")

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert 'filename="donnees_budgetpilot360_' in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == [
        "Entreprise", "Équipe", "Catégories", "Dépenses", "Recettes",
        "Automatisations", "Journal d'audit",
    ]

    entreprise = wb["Entreprise"]
    assert "Acme" in entreprise["A1"].value
    assert entreprise["B3"].value == "Acme"
    assert entreprise["B9"].value == "1 / 2"  # 1 actif (Awa), Jean retiré

    equipe = wb["Équipe"]
    assert equipe.max_row == 3
    assert equipe["D2"].value == "Admin principal"  # a1 = owner_id
    assert equipe["E3"].value == "Retiré"

    depenses = wb["Dépenses"]
    assert depenses["B2"].value == "Transport"
    assert depenses["F2"].value == "Approuvée"
    assert depenses["H2"].value == "Oui"  # justificatif présent
    assert depenses["J2"].value == "Awa Admin"  # validée par

    recettes = wb["Recettes"]
    assert recettes["D2"].value == "Client A"
    assert recettes["G2"].value == "Confirmée"
    assert recettes["J2"].value == "Oui"  # issue d'une automatisation

    autos = wb["Automatisations"]
    assert autos["A2"].value == "Dépense"
    assert autos["F2"].value == "2/12"

    audit_ws = wb["Journal d'audit"]
    assert audit_ws["B2"].value == "expense.approved"
    assert audit_ws["C2"].value == "Awa Admin"

    entry = next(c for c in silence_audit if c["action"] == "company.data_exported")
    assert entry["company_id"] == as_admin.company_id


def test_company_export_forbidden_for_user(client, as_user, monkeypatch):
    mock_client = _mock_full_export_client(monkeypatch)

    assert client.get("/reports/company-export").status_code == 403
    mock_client.table.assert_not_called()


def test_company_export_requires_company(client, as_super_admin, monkeypatch):
    """Le super_admin Pukri n'a pas d'entreprise : pas d'export global par ce biais."""
    _mock_full_export_client(monkeypatch)

    assert client.get("/reports/company-export").status_code == 400
